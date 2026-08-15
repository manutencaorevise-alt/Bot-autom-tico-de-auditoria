"""
Spreadsheet Handler - Lê histórico da planilha e registra novos orçamentos
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from copy import copy
from datetime import datetime
from typing import Dict, List, Optional
import logging

logger = logging.getLogger(__name__)

class SpreadsheetHandler:
    def __init__(self, filepath: str):
        """Abre planilha BASE_E_HISTORICO_ATUALIZADA.xlsx"""
        try:
            self.wb = openpyxl.load_workbook(filepath)
            self.filepath = filepath
            logger.info(f"Planilha carregada: {filepath}")
        except FileNotFoundError:
            logger.error(f"Planilha não encontrada: {filepath}")
            self.wb = None

    def get_vehicle_history(self, placa: str) -> List[Dict]:
        """Busca histórico completo de uma placa na Base_Itens"""
        if not self.wb:
            return []

        ws = self.wb['Base_Itens']
        history = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            # Coluna 7 = Placa (index 6)
            cell_placa = row[6]
            if cell_placa and cell_placa.value:
                if str(cell_placa.value).upper().strip() == placa.upper().strip():
                    # Extrai dados da linha
                    row_data = {
                        'data': row[0].value,
                        'empresa': row[1].value,
                        'cidade': row[2].value,
                        'uf': row[3].value,
                        'licitation': row[4].value,
                        'orcamento': row[5].value,
                        'placa': row[6].value,
                        'veiculo': row[7].value,
                        'grupo': row[8].value,
                        'km': row[9].value,
                        'categoria': row[10].value,
                        'item': row[11].value,
                        'item_normalizado': row[12].value,
                        'qtd': row[13].value,
                        'valor_unitario': row[14].value,
                        'valor_total': row[15].value,
                        'total_orcamento': row[16].value,
                        'observacoes': row[21].value,
                    }
                    history.append(row_data)

        logger.info(f"Histórico encontrado para {placa}: {len(history)} registros")
        return history

    def get_oficina_history(self, oficina_name: str) -> List[Dict]:
        """Busca histórico de todos os orçamentos de uma oficina"""
        if not self.wb:
            return []

        ws = self.wb['Base_Itens']
        history = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            # Coluna 2 = Empresa (index 1)
            cell_empresa = row[1]
            if cell_empresa and cell_empresa.value:
                if oficina_name.lower() in str(cell_empresa.value).lower():
                    row_data = {
                        'data': row[0].value,
                        'empresa': row[1].value,
                        'placa': row[6].value,
                        'veiculo': row[7].value,
                        'item': row[11].value,
                        'valor_unitario': row[14].value,
                        'valor_total': row[15].value,
                    }
                    history.append(row_data)

        logger.info(f"Histórico da oficina '{oficina_name}': {len(history)} registros")
        return history

    def register_new_budget(self, budget_data: Dict) -> bool:
        """Registra novo orçamento em Orcamentos_Recentes"""
        if not self.wb:
            logger.error("Planilha não carregada")
            return False

        try:
            ws = self.wb['Orcamentos_Recentes']
            last_row = ws.max_row
            new_row = last_row + 1

            # Copia formatação da última linha
            template_row = last_row
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=template_row, column=col)
                dst_cell = ws.cell(row=new_row, column=col)

                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

            # Insere dados do orçamento
            ws.cell(row=new_row, column=1, value=datetime.now())
            ws.cell(row=new_row, column=2, value=budget_data.get('orcamento'))
            ws.cell(row=new_row, column=3, value=budget_data.get('placa'))
            ws.cell(row=new_row, column=4, value=budget_data.get('veiculo'))
            ws.cell(row=new_row, column=5, value=budget_data.get('oficina'))
            ws.cell(row=new_row, column=6, value=budget_data.get('cidade_uf'))
            ws.cell(row=new_row, column=7, value=budget_data.get('km'))
            ws.cell(row=new_row, column=8, value=budget_data.get('valor_total'))
            ws.cell(row=new_row, column=9, value=budget_data.get('status'))
            ws.cell(row=new_row, column=10, value=budget_data.get('acima_alcada'))
            ws.cell(row=new_row, column=11, value=budget_data.get('observacao'))
            ws.cell(row=new_row, column=12, value=budget_data.get('arquivo'))

            self.wb.save(self.filepath)
            logger.info(f"Orçamento registrado em Orcamentos_Recentes (linha {new_row})")
            return True

        except Exception as e:
            logger.error(f"Erro registrando orçamento: {e}")
            return False

    def register_new_item(self, item_data: Dict) -> bool:
        """Registra novo item em Base_Itens"""
        if not self.wb:
            logger.error("Planilha não carregada")
            return False

        try:
            ws = self.wb['Base_Itens']
            last_row = ws.max_row
            new_row = last_row + 1

            # Copia formatação
            template_row = last_row
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=template_row, column=col)
                dst_cell = ws.cell(row=new_row, column=col)

                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

            # Insere dados
            ws.cell(row=new_row, column=1, value=item_data.get('data', datetime.now()))
            ws.cell(row=new_row, column=2, value=item_data.get('empresa'))
            ws.cell(row=new_row, column=3, value=item_data.get('cidade'))
            ws.cell(row=new_row, column=4, value=item_data.get('uf'))
            ws.cell(row=new_row, column=5, value=item_data.get('licitacao'))
            ws.cell(row=new_row, column=6, value=item_data.get('orcamento'))
            ws.cell(row=new_row, column=7, value=item_data.get('placa'))
            ws.cell(row=new_row, column=8, value=item_data.get('veiculo'))
            ws.cell(row=new_row, column=9, value=item_data.get('grupo'))
            ws.cell(row=new_row, column=10, value=item_data.get('km'))
            ws.cell(row=new_row, column=11, value=item_data.get('categoria'))
            ws.cell(row=new_row, column=12, value=item_data.get('item'))
            ws.cell(row=new_row, column=13, value=item_data.get('item_normalizado'))
            ws.cell(row=new_row, column=14, value=item_data.get('qtd'))
            ws.cell(row=new_row, column=15, value=item_data.get('valor_unitario'))
            ws.cell(row=new_row, column=16, value=item_data.get('valor_total'))
            ws.cell(row=new_row, column=17, value=item_data.get('total_orcamento'))
            ws.cell(row=new_row, column=18, value=item_data.get('acima_alcada'))
            ws.cell(row=new_row, column=19, value=item_data.get('decisao_padrao'))
            ws.cell(row=new_row, column=20, value=item_data.get('decisao_detalhe'))
            ws.cell(row=new_row, column=21, value=item_data.get('fonte'))
            ws.cell(row=new_row, column=22, value=item_data.get('observacoes'))
            ws.cell(row=new_row, column=23, value=item_data.get('arquivos'))

            self.wb.save(self.filepath)
            logger.info(f"Item registrado em Base_Itens (linha {new_row})")
            return True

        except Exception as e:
            logger.error(f"Erro registrando item: {e}")
            return False

    def get_similar_services(self, placa: str, service_keyword: str) -> List[Dict]:
        """Busca serviços similares anteriores da mesma placa"""
        if not self.wb:
            return []

        history = self.get_vehicle_history(placa)
        similar = []

        for record in history:
            if record.get('item') and service_keyword.lower() in record.get('item', '').lower():
                similar.append(record)

        logger.info(f"Serviços similares para {placa} ({service_keyword}): {len(similar)}")
        return similar
